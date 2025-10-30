from django.contrib import admin
from .models import Book, Borrow, Notification, EntryLog, Collection, SubCollection
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook
from django.template.response import TemplateResponse
from datetime import date
from django.urls import path
from django.shortcuts import redirect
from django.utils.safestring import mark_safe
from openpyxl.utils import get_column_letter

@admin.register(Collection)
class CollectionAdmin(admin.ModelAdmin):
    list_display = ('name', 'description')
    search_fields = ('name',)
    list_per_page = 10

@admin.register(SubCollection)
class SubCollectionAdmin(admin.ModelAdmin):
    list_display = ('name', 'collection')
    list_filter = ('collection',)
    search_fields = ('name', 'collection__name')
    list_per_page = 10


@admin.register(Book)
class BookAdmin(admin.ModelAdmin):
    list_display = ('title', 'author', 'quantity', 'get_collection', 'subcollection', 'publish_year', 'publisher')
    search_fields = ("title", "author")
    list_filter = ('subcollection__collection', 'subcollection')
    list_per_page = 10 #tuy chinh phan trang
    autocomplete_fields = ('subcollection',)

    def get_collection(self, obj):
        return obj.subcollection.collection.name if obj.subcollection else "-"
    get_collection.short_description = "Bộ sưu tập"

class DueDateFilter(admin.SimpleListFilter):
    """Bộ lọc tùy chọn theo hạn trả."""
    title = "Hạn trả"
    parameter_name = "due_range"

    def lookups(self, request, model_admin):
        return [
            ("today", "Hôm nay"),
            ("7days", "7 ngày tới"),
            ("thismonth", "Tháng này"),
            ("nextmonth", "Tháng tới"),
        ]

    def queryset(self, request, queryset):
        today = timezone.now().date()
        if self.value() == "today":
            return queryset.filter(due_date=today)
        elif self.value() == "7days":
            return queryset.filter(due_date__range=(today, today + timedelta(days=7)))
        elif self.value() == "thismonth":
            return queryset.filter(due_date__year=today.year, due_date__month=today.month)
        elif self.value() == "nextmonth":
            next_month = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
            return queryset.filter(due_date__year=next_month.year, due_date__month=next_month.month)
        return queryset

@admin.register(Borrow)
class BorrowAdmin(admin.ModelAdmin):
    list_display = (
        "borrow_code", "user", "book", "formatted_borrow_date",
        "formatted_due_date", "formatted_return_date", "status"
    )
    list_filter = ("status", "borrow_date", DueDateFilter)
    search_fields = ("borrow_code", "user__username", "book__title")
    ordering = ("-borrow_date",)
    list_per_page = 10 #tuy chinh phan trang

    # Ghi đè template trang danh sách để chèn nút thống kê
    change_list_template = "admin/borrow_change_list.html"

    def formatted_borrow_date(self, obj):
        return obj.borrow_date.strftime("%d/%m/%Y") if obj.borrow_date else ""
    formatted_borrow_date.short_description = "Ngày mượn"

    def formatted_due_date(self, obj):
        return obj.due_date.strftime("%d/%m/%Y") if obj.due_date else ""
    formatted_due_date.short_description = "Hạn trả"

    def formatted_return_date(self, obj):
        return obj.return_date.strftime("%d/%m/%Y") if obj.return_date else ""
    formatted_return_date.short_description = "Ngày trả"

    # ==== Custom URL cho trang thống kê ====
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("statistics/", self.admin_site.admin_view(self.borrow_statistics_view),
                 name="borrow_statistics"),
            path("export_overdue/", self.admin_site.admin_view(self.export_overdue_to_excel),
                 name="export_overdue_to_excel"),
        ]
        return custom_urls + urls

    # ==== Trang thống kê ====
    def borrow_statistics_view(self, request):
        top_books = (
            Borrow.objects.filter(status="Đang mượn")
            .values("book__title")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )

        overdue_borrows = Borrow.objects.filter(status="Đang mượn", due_date__lt=date.today())

        context = dict(
            self.admin_site.each_context(request),
            title="Thống kê mượn sách",
            top_books=top_books,
            overdue_borrows=overdue_borrows,
        )

        return TemplateResponse(request, "admin/borrow_statistics.html", context)

    # ==== Xuất Excel ====
    def export_overdue_to_excel(self, request):
        overdue_borrows = Borrow.objects.filter(status="Đang mượn", due_date__lt=date.today())

        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách quá hạn"

        headers = ["Mã mượn", "Người mượn", "Tên sách", "Tác giả", "Ngày mượn", "Hạn trả"]
        ws.append(headers)

        for b in overdue_borrows:
            ws.append([
                b.borrow_code,
                b.user.username,
                b.book.title,
                b.book.author,
                b.borrow_date.strftime("%d/%m/%Y") if b.borrow_date else "",
                b.due_date.strftime("%d/%m/%Y") if b.due_date else "",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="overdue_borrows.xlsx"'
        # Căn chỉnh độ rộng cột tự động
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = (length + 2)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        # Tăng nhẹ chiều cao các hàng để dễ đọc
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 20
        wb.save(response)
        return response
    
@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ("user", "title", "is_read", "formatted_created_at")
    list_filter = ("is_read", "created_at")
    search_fields = ("user__username", "title", "message")
    ordering = ("-created_at",)

    def formatted_created_at(self, obj):
        if obj.created_at:
            local_time = timezone.localtime(obj.created_at) 
            return local_time.strftime("%H:%M %d/%m/%Y")
        return "-"
    formatted_created_at.short_description = "Thời gian tạo"

@admin.register(EntryLog)
class EntryLogAdmin(admin.ModelAdmin):
    list_display = ('user', 'shift', 'formatted_check_in', 'formatted_check_out')
    list_filter = ('shift', 'check_in', 'check_out')
    search_fields = ('user__username',)
    ordering = ('-check_in',)
    date_hierarchy = 'check_in'
    list_per_page = 20

    def formatted_check_in(self, obj):
        if obj.check_in:
            return timezone.localtime(obj.check_in).strftime("%H:%M %d/%m/%Y")
        return "-"
    formatted_check_in.short_description = "Giờ vào"

    def formatted_check_out(self, obj):
        if obj.check_out:
            return timezone.localtime(obj.check_out).strftime("%H:%M %d/%m/%Y")
        return "-"
    formatted_check_out.short_description = "Giờ ra"

